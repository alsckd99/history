import requests
import json
import re
import csv
import os
from typing import Optional, Dict, List, Any
from pathlib import Path
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# 엑셀 처리를 위한 라이브러리 임포트 시도
try:
    import openpyxl
except ImportError:
    openpyxl = None

def clean_footnote(footnote_text: str) -> List[Dict[str, str]]:
    """
    본문 주석 설명을 LLM이 해석하기 좋은 구조화된 형식으로 변환합니다.
    
    Returns:
        list: [{"id": "1", "term": "훈척", "definition": "설명..."}, ...]
    """
    if not footnote_text:
        return []
    
    # \r\n을 \n으로 통일
    text = footnote_text.replace('\r\n', '\n')
    
    # [^1]: 패턴으로 분리
    pattern = r'\[\^(\d+)\]:\s*'
    parts = re.split(pattern, text)
    
    annotations = []
    # parts는 ['', '1', '내용1', '2', '내용2', ...] 형태
    i = 1
    while i < len(parts) - 1:
        note_id = parts[i]
        content = parts[i + 1].strip()
        
        if content:
            # [→우리말샘](URL) 패턴 제거
            content = re.sub(r'\s*\[→[^\]]+\]\([^)]+\)', '', content)
            
            # 용어와 설명 분리 (첫 번째 문장이나 콜론으로 구분)
            # 예: "훈척(勳戚) : 대대로 나라나..." 또는 "훈척. 대대로..."
            term = ""
            definition = content
            
            # 콜론이나 마침표로 분리 시도
            if ': ' in content[:50]:
                parts_split = content.split(': ', 1)
                term = parts_split[0].strip()
                definition = parts_split[1].strip() if len(parts_split) > 1 else content
            elif '. ' in content[:50]:
                parts_split = content.split('. ', 1)
                term = parts_split[0].strip()
                definition = parts_split[1].strip() if len(parts_split) > 1 else content
            
            annotations.append({
                "id": note_id,
                "term": term,
                "definition": definition
            })
        
        i += 2
    
    return annotations


def html_table_to_structured(html_content: str) -> Dict[str, Any]:
    """
    HTML 콘텐츠에서 테이블을 구조화된 데이터로 추출하고,
    나머지 본문은 텍스트로 반환합니다.
    """
    if not html_content:
        return {"text": "", "tables": []}
    
    # \r\n을 실제 줄바꿈으로 변환
    text = html_content.replace('\\r\\n', '\n').replace('\r\n', '\n')
    
    # <table> 태그가 있는지 확인
    if '<table' not in text.lower():
        # HTML 태그 제거
        soup = BeautifulSoup(text, 'html.parser')
        clean_text = soup.get_text(separator='\n')
        clean_text = re.sub(r'\n{3,}', '\n\n', clean_text)
        return {"text": clean_text.strip(), "tables": []}
    
    soup = BeautifulSoup(text, 'html.parser')
    tables_data = []
    
    # 테이블 처리
    for table in soup.find_all('table'):
        table_info = {"title": "", "headers": [], "rows": []}
        
        # tfoot에서 테이블 제목 추출
        tfoot = table.find('tfoot')
        if tfoot:
            table_info["title"] = tfoot.get_text(strip=True)
        
        # thead에서 헤더 추출
        headers = []
        thead = table.find('thead')
        if thead:
            # 마지막 헤더 행에서 열 이름 추출
            header_rows = thead.find_all('tr')
            if header_rows:
                last_header = header_rows[-1]
                cells = last_header.find_all(['th', 'td'])
                headers = [cell.get_text(strip=True) for cell in cells]
        
        # 헤더가 없으면 첫 번째 행에서 추출 시도
        if not headers:
            first_row = table.find('tr')
            if first_row:
                cells = first_row.find_all(['th', 'td'])
                headers = [cell.get_text(strip=True) for cell in cells]
        
        table_info["headers"] = headers
        
        # tbody에서 데이터 행 추출
        tbody = table.find('tbody')
        rows_to_process = tbody.find_all('tr') if tbody else table.find_all('tr')[1:]
        
        for row in rows_to_process:
            cells = row.find_all(['th', 'td'])
            cell_values = [cell.get_text(strip=True) for cell in cells]
            
            # 헤더와 매핑하여 딕셔너리로 변환
            if headers and len(cell_values) == len(headers):
                row_dict = dict(zip(headers, cell_values))
            else:
                # 헤더 수와 맞지 않으면 인덱스로 매핑
                row_dict = {f"col_{i}": v for i, v in enumerate(cell_values)}
            
            table_info["rows"].append(row_dict)
        
        tables_data.append(table_info)
        
        # 테이블 자리에 표시자 남기기
        table.replace_with(f'[표: {table_info["title"]}]')
    
    # 나머지 HTML 태그 제거하고 텍스트만 추출
    result_text = soup.get_text(separator='\n')
    result_text = re.sub(r'\n{3,}', '\n\n', result_text)
    
    return {"text": result_text.strip(), "tables": tables_data}


def convert_related_articles(related_articles: List[Dict]) -> List[Dict]:
    """
    관련항목의 필드명을 한글로 변환하고 writerInfo 제거
    """
    if not related_articles:
        return []
    
    converted = []
    for article in related_articles:
        converted_article = {
            'URL': article.get('url', ''),
            '항목명': article.get('headword', ''),
            '원어': article.get('origin', ''),
            '항목 분야': article.get('field', ''),
            '항목 유형': article.get('contentsType', ''),
            '시대': article.get('era', ''),
            '항목 정의': article.get('definition', '')
            # writerInfo 제거
        }
        converted.append(converted_article)
    
    return converted

class HistoryAPIClient:
    """한국역사정보통합시스템 API 클라이언트"""
    
    def __init__(self, api_key: str):
        self.base_url = "https://devin.aks.ac.kr:8080/v1"
        self.headers = {
            "X-API-Key": api_key
        }
    
    def get_all_articles(self, page_no: int = 1) -> Dict[str, Any]:
        url = f"{self.base_url}/Articles"
        params = {"pageNo": page_no}
        response = requests.get(url, headers=self.headers, params=params)
        print(f"[요청 URL] {response.url}")
        response.raise_for_status()
        return response.json()
    
    def search_articles(self, keyword: str, page: int = 1, field: Optional[str] = None) -> Dict[str, Any]:
        url = f"{self.base_url}/Articles/Search"
        params = {"keyword": keyword, "page": page}
        if field:
            params["field"] = field
        response = requests.get(url, headers=self.headers, params=params)
        print(f"[요청 URL] {response.url}")
        response.raise_for_status()
        return response.json()
    
    def get_articles_by_field(self, field: str, page_no: int = 1) -> Dict[str, Any]:
        url = f"{self.base_url}/Articles/Field/{field}"
        params = {"pageNo": page_no}
        response = requests.get(url, headers=self.headers, params=params)
        print(f"[요청 URL] {response.url}")
        response.raise_for_status()
        return response.json()
    
    def get_category_fields(self) -> Dict[str, Any]:
        url = f"{self.base_url}/Category/Field"
        response = requests.get(url, headers=self.headers)
        print(f"[요청 URL] {response.url}")
        response.raise_for_status()
        return response.json()
    
    def get_category_contents_types(self) -> Dict[str, Any]:
        """항목 유형(contentsType) 카테고리 조회"""
        url = f"{self.base_url}/Articles/ContentsType"
        response = requests.get(url, headers=self.headers)
        print(f"[요청 URL] {response.url}")
        response.raise_for_status()
        return response.json()
    
    def get_article_detail(self, article_id: str) -> Dict[str, Any]:
        url = f"{self.base_url}/Article/{article_id}"
        response = requests.get(url, headers=self.headers)
        # print(f"[요청 URL] {response.url}") # 너무 많은 로그 방지
        response.raise_for_status()
        return response.json()
    
    def search_medias(self, keyword: str, page_no: int = 1) -> Dict[str, Any]:
        url = f"{self.base_url}/Medias/Search"
        params = {"keyword": keyword, "pageNo": page_no}
        response = requests.get(url, headers=self.headers, params=params)
        print(f"[요청 URL] {response.url}")
        response.raise_for_status()
        return response.json()
    
    def get_media_detail(self, media_id: str) -> Dict[str, Any]:
        url = f"{self.base_url}/Media/{media_id}"
        response = requests.get(url, headers=self.headers)
        print(f"[요청 URL] {response.url}")
        response.raise_for_status()
        return response.json()
    
    def _fetch_single_article(self, article_id: str) -> Optional[Dict]:
        """단일 항목 조회 (병렬 처리용 내부 메서드)"""
        try:
            response_data = self.get_article_detail(article_id)
            article = response_data.get('article', {})
            
            # 본문에서 HTML 테이블을 구조화된 데이터로 변환
            body_data = html_table_to_structured(article.get('body', ''))
            
            # 본문 주석을 구조화된 형식으로 변환
            footnotes = clean_footnote(article.get('footNote', ''))
            
            # 관련항목 필드명 변환 및 writerInfo 제거
            related = convert_related_articles(article.get('relatedArticles', []))
            
            article_info = {
                'url': article.get('url'),
                '항목명': article.get('headword'),
                '원어': article.get('origin'),
                '항목 분야': article.get('field'),
                '항목 유형': article.get('contentsType'),
                '시대': article.get('era'),
                '항목 정의': article.get('definition'),
                '요약': article.get('summary'),
                '키워드': article.get('keyword'),
                '항목 본문': body_data.get('text', ''),
                '본문 표': body_data.get('tables', []),
                '주석': footnotes,
                '관련항목': related
            }
            
            return article_info
            
        except requests.exceptions.RequestException as e:
            print(f"\n❌ {article_id} 조회 실패: {e}")
            return None
        except Exception as e:
            print(f"\n❌ {article_id} 처리 중 오류 발생: {e}")
            return None

    def save_articles_to_json(
        self, 
        article_ids: List[str], 
        output_file: str = "articles_data.json",
        max_workers: int = 12,
        use_parallel: bool = True
    ):
        """
        여러 항목의 상세 정보를 조회하여 JSON 파일로 저장
        
        Args:
            article_ids: 조회할 항목 ID 리스트
            output_file: 저장할 JSON 파일 경로
            max_workers: 병렬 처리 시 최대 스레드 수 (기본: 10)
            use_parallel: 병렬 처리 사용 여부 (기본: True)
        """
        articles_data = []
        
        # 중복 ID 제거
        unique_ids = list(set(article_ids))
        total_count = len(unique_ids)
        print(f"총 {total_count}개의 고유 ID에 대해 작업을 시작합니다.")
        
        if use_parallel and total_count > 1:
            # ===== 병렬 처리 =====
            print(f"🚀 병렬 처리 모드 (스레드: {max_workers}개)")
            
            completed_count = 0
            failed_count = 0
            lock = threading.Lock()
            
            def process_with_progress(article_id: str) -> Optional[Dict]:
                nonlocal completed_count, failed_count
                result = self._fetch_single_article(article_id)
                
                with lock:
                    if result:
                        completed_count += 1
                    else:
                        failed_count += 1
                    # 진행률 표시
                    print(f"\r[{completed_count + failed_count}/{total_count}] "
                          f"완료: {completed_count} | 실패: {failed_count}", end='')
                
                return result
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # 모든 작업 제출
                futures = {
                    executor.submit(process_with_progress, article_id): article_id 
                    for article_id in unique_ids
                }
                
                # 완료된 순서대로 결과 수집
                for future in as_completed(futures):
                    result = future.result()
                    if result:
                        articles_data.append(result)
            
            print()  # 줄바꿈
            
        else:
            # ===== 순차 처리 =====
            print("📝 순차 처리 모드")
            
            for i, article_id in enumerate(unique_ids, 1):
                print(f"[{i}/{total_count}] 항목 조회 중: {article_id}", end='\r')
                
                result = self._fetch_single_article(article_id)
                if result:
                    articles_data.append(result)
        
        # JSON 파일로 저장
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(articles_data, f, ensure_ascii=False, indent=2)
        
        print(f"\n✅ 총 {len(articles_data)}개 항목을 {output_file}에 저장했습니다.")
        return articles_data
    
    def save_all_articles_from_page(self, page_no: int = 1, output_file: str = None):
        if output_file is None:
            output_file = f"articles_page_{page_no}.json"
        
        print(f"=== 페이지 {page_no} 항목 리스트 조회 중 ===")
        all_articles = self.get_all_articles(page_no=page_no)
        
        article_ids = [article['eid'] for article in all_articles.get('articles', [])]
        print(f"총 {len(article_ids)}개 항목을 조회합니다.\n")
        
        return self.save_articles_to_json(article_ids, output_file)

def _parse_id_from_url(url: str) -> Optional[str]:
    """URL에서 ID를 파싱하는 내부 헬퍼 함수"""
    if not url:
        return None
        
    try:
        # 예: http://encykorea.aks.ac.kr/Contents/Item/E0000072
        extracted_id = url.split('/')[-1]
        
        # ID가 'E'로 시작하는지 확인
        if extracted_id.startswith('E'):
            return extracted_id
        else:
            # 쿼리 파라미터 처리 (예: E0000072?view=...)
            match = re.search(r'(E\d+)', extracted_id)
            if match:
                return match.group(1)
    except Exception:
        pass
    return None

def extract_ids_from_excel(file_path: str) -> List[str]:
    """
    Excel(.xlsx) 파일을 읽어 C열(3번째 열)의 URL에서 ID를 추출합니다.
    """
    if openpyxl is None:
        print("❌ 'openpyxl' 라이브러리가 설치되어 있지 않습니다.")
        print("   설치 명령어: pip install openpyxl")
        return []

    ids = []
    if not os.path.exists(file_path):
        print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
        return []

    try:
        print(f"📗 엑셀 파일 열기: {file_path}")
        # data_only=True: 수식이 아닌 값만 읽음
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet = wb.active # 첫 번째 시트 활성화
        
        # iter_rows를 사용하여 메모리 효율적으로 읽기
        # min_row=2 (헤더 건너뛰기), min_col=3, max_col=3 (C열만)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            if row and row[0]:
                url = str(row[0]).strip()
                article_id = _parse_id_from_url(url)
                if article_id:
                    ids.append(article_id)
        
        wb.close()
        print(f"✅ 엑셀 파일 읽기 성공")
        
    except Exception as e:
        print(f"❌ 엑셀 파일 처리 중 오류 발생: {e}")
        
    return ids

def extract_ids_from_csv(file_path: str, encoding: str = None) -> List[str]:
    """
    CSV 파일을 읽어 C열(3번째 열)의 URL에서 ID를 추출합니다.
    """
    ids = []
    if not os.path.exists(file_path):
        print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
        return []

    encodings_to_try = ['utf-8', 'utf-8-sig', 'cp949', 'euc-kr']
    if encoding:
        encodings_to_try = [encoding]

    for enc in encodings_to_try:
        try:
            temp_ids = []
            with open(file_path, 'r', encoding=enc) as f:
                reader = csv.reader(f)
                headers = next(reader, None) # 헤더 스킵
                
                for row in reader:
                    if len(row) > 2:
                        url = row[2].strip()
                        article_id = _parse_id_from_url(url)
                        if article_id:
                            temp_ids.append(article_id)
            
            print(f"✅ 인코딩 감지 성공: '{enc}'")
            return temp_ids

        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"❌ '{enc}' 시도 중 읽기 오류 발생: {e}")
            continue
            
    print(f"❌ 모든 인코딩으로 파일을 읽을 수 없습니다.")
    return []

def extract_ids_from_file(file_path: str) -> List[str]:
    """파일 확장자에 따라 적절한 추출 함수를 호출합니다."""
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.xlsx':
        return extract_ids_from_excel(file_path)
    elif ext == '.csv':
        return extract_ids_from_csv(file_path)
    else:
        print(f"❌ 지원하지 않는 파일 형식입니다: {ext}")
        return []

# 사용 예시
if __name__ == "__main__":
    # API 키 설정
    API_KEY = "A5931F50-59E2-4679-93CF-5858EC174900"
    
    # 클라이언트 생성
    client = HistoryAPIClient(API_KEY)
    
    # ==========================================
    # 파일 설정 (여기를 수정하세요)
    # .csv 또는 .xlsx 파일 경로 입력
    # ==========================================
    # 예: "data.xlsx" 또는 "data.csv"
    input_file_path = "한국학중앙연구원_한국민족문화대백과사전_20240130.xlsx" 
    output_json_path = "crawling/results.json"
    
    print(f"=== 파일 처리 시작: {input_file_path} ===")
    
    # 카테고리 필드 조회
    fields = client.get_category_fields()
    print("\n=== 카테고리 필드 목록 ===")
    print(json.dumps(fields, ensure_ascii=False, indent=2))
    
    # 항목 유형(contentsType) 조회
    contents_types = client.get_category_contents_types()
    print("\n=== 항목 유형(ContentsType) 목록 ===")
    print(json.dumps(contents_types, ensure_ascii=False, indent=2))
    
    # # 1. 파일에서 ID 추출 (자동 감지)
    # target_ids = extract_ids_from_file(input_file_path)
    
    # if target_ids:
    #     print(f"\n🔍 추출된 ID 개수: {len(target_ids)}개")
    #     print(f"   (첫 5개 예시: {target_ids[:5]})")
        
    #     # 2. 추출된 ID로 API 조회 및 저장
    #     print("\n=== API 데이터 수집 시작 ===")
    #     client.save_articles_to_json(target_ids, output_json_path)
    # else:
    #     print("\n⚠️ 처리할 ID가 없습니다. 파일 경로와 내용을 확인해주세요.")